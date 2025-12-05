"""
Data Export Module - Export data in multiple formats

Supports exporting:
- Entire databases
- Individual schemas
- Selected tables
- Query results
- DDL statements

Export formats:
- CSV
- JSON  
- Parquet
- SQL (DDL + DML)
- Excel
"""

import os
import io
import csv
import json
import logging
import zipfile
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Union, BinaryIO
from pathlib import Path

logger = logging.getLogger("snowglobe.data_export")


class DataExporter:
    """
    Export data from Snowglobe in various formats.
    """
    
    SUPPORTED_FORMATS = ['csv', 'json', 'parquet', 'sql', 'excel', 'jsonl']
    
    def __init__(self, query_executor):
        """
        Initialize the data exporter.
        
        Args:
            query_executor: QueryExecutor instance for database operations
        """
        self.executor = query_executor
        self.metadata = query_executor.metadata
    
    # ==================== Query Result Export ====================
    
    def export_query_result(self, sql: str, format: str = 'csv', 
                            options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Export query results to a specific format.
        
        Args:
            sql: SQL query to execute
            format: Output format (csv, json, parquet, excel)
            options: Export options
            
        Returns:
            Dictionary with export results (content or file path)
        """
        options = options or {}
        format = format.lower()
        
        if format not in self.SUPPORTED_FORMATS:
            return {
                "success": False,
                "error": f"Unsupported format: {format}. Supported: {', '.join(self.SUPPORTED_FORMATS)}"
            }
        
        # Execute the query
        result = self.executor.execute(sql)
        if not result["success"]:
            return {
                "success": False,
                "error": result.get("error", "Query execution failed")
            }
        
        columns = result["columns"]
        data = result["data"]
        
        return self._export_data(columns, data, format, options)
    
    def _export_data(self, columns: List[str], data: List[List], 
                     format: str, options: Dict[str, Any]) -> Dict[str, Any]:
        """
        Export data in the specified format.
        
        Args:
            columns: Column names
            data: Row data
            format: Export format
            options: Format-specific options
            
        Returns:
            Export result dictionary
        """
        try:
            if format == 'csv':
                return self._export_csv(columns, data, options)
            elif format == 'json':
                return self._export_json(columns, data, options)
            elif format == 'jsonl':
                return self._export_jsonl(columns, data, options)
            elif format == 'parquet':
                return self._export_parquet(columns, data, options)
            elif format == 'excel':
                return self._export_excel(columns, data, options)
            elif format == 'sql':
                return self._export_sql_insert(columns, data, options)
            else:
                return {"success": False, "error": f"Unknown format: {format}"}
        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
    
    def _export_csv(self, columns: List[str], data: List[List], 
                    options: Dict[str, Any]) -> Dict[str, Any]:
        """Export to CSV format."""
        delimiter = options.get('delimiter', ',')
        quote_char = options.get('quote_char', '"')
        include_header = options.get('include_header', True)
        null_value = options.get('null_value', '')
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter, quotechar=quote_char, 
                           quoting=csv.QUOTE_MINIMAL)
        
        if include_header:
            writer.writerow(columns)
        
        for row in data:
            writer.writerow([null_value if v is None else v for v in row])
        
        content = output.getvalue()
        
        return {
            "success": True,
            "format": "csv",
            "content": content,
            "content_type": "text/csv",
            "filename": f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "row_count": len(data),
            "column_count": len(columns)
        }
    
    def _export_json(self, columns: List[str], data: List[List], 
                     options: Dict[str, Any]) -> Dict[str, Any]:
        """Export to JSON format."""
        orient = options.get('orient', 'records')  # records, columns, values
        indent = options.get('indent', 2)
        
        if orient == 'records':
            # Each row as a dictionary
            json_data = [dict(zip(columns, row)) for row in data]
        elif orient == 'columns':
            # Each column as a list
            json_data = {col: [row[i] for row in data] for i, col in enumerate(columns)}
        elif orient == 'values':
            # Just the data array with schema
            json_data = {
                "columns": columns,
                "data": data
            }
        else:
            json_data = [dict(zip(columns, row)) for row in data]
        
        content = json.dumps(json_data, indent=indent, default=str)
        
        return {
            "success": True,
            "format": "json",
            "content": content,
            "content_type": "application/json",
            "filename": f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "row_count": len(data),
            "column_count": len(columns)
        }
    
    def _export_jsonl(self, columns: List[str], data: List[List], 
                      options: Dict[str, Any]) -> Dict[str, Any]:
        """Export to JSON Lines format (one JSON object per line)."""
        lines = []
        for row in data:
            obj = dict(zip(columns, row))
            lines.append(json.dumps(obj, default=str))
        
        content = '\n'.join(lines)
        
        return {
            "success": True,
            "format": "jsonl",
            "content": content,
            "content_type": "application/x-ndjson",
            "filename": f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jsonl",
            "row_count": len(data),
            "column_count": len(columns)
        }
    
    def _export_parquet(self, columns: List[str], data: List[List], 
                        options: Dict[str, Any]) -> Dict[str, Any]:
        """Export to Parquet format."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            return {
                "success": False,
                "error": "pyarrow is required for Parquet export. Install with: pip install pyarrow"
            }
        
        # Create PyArrow table
        table_dict = {col: [row[i] for row in data] for i, col in enumerate(columns)}
        table = pa.table(table_dict)
        
        # Write to bytes
        output = io.BytesIO()
        pq.write_table(table, output, compression=options.get('compression', 'snappy'))
        output.seek(0)
        
        return {
            "success": True,
            "format": "parquet",
            "content": output.getvalue(),
            "content_type": "application/octet-stream",
            "filename": f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.parquet",
            "row_count": len(data),
            "column_count": len(columns),
            "binary": True
        }
    
    def _export_excel(self, columns: List[str], data: List[List], 
                      options: Dict[str, Any]) -> Dict[str, Any]:
        """Export to Excel format."""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            # Fallback to CSV-like format
            return self._export_csv(columns, data, options)
        
        sheet_name = options.get('sheet_name', 'Data')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write header
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col)
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "success": True,
            "format": "excel",
            "content": output.getvalue(),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "row_count": len(data),
            "column_count": len(columns),
            "binary": True
        }
    
    def _export_sql_insert(self, columns: List[str], data: List[List], 
                           options: Dict[str, Any]) -> Dict[str, Any]:
        """Export as SQL INSERT statements."""
        table_name = options.get('table_name', 'exported_table')
        batch_size = options.get('batch_size', 1000)
        include_create = options.get('include_create', False)
        
        lines = []
        
        # Add CREATE TABLE if requested
        if include_create:
            col_defs = ", ".join(f"{col} VARCHAR" for col in columns)
            lines.append(f"CREATE TABLE IF NOT EXISTS {table_name} ({col_defs});")
            lines.append("")
        
        # Generate INSERT statements
        col_list = ", ".join(columns)
        
        for i, row in enumerate(data):
            values = []
            for v in row:
                if v is None:
                    values.append("NULL")
                elif isinstance(v, str):
                    escaped = v.replace("'", "''")
                    values.append(f"'{escaped}'")
                elif isinstance(v, bool):
                    values.append("TRUE" if v else "FALSE")
                else:
                    values.append(str(v))
            
            values_str = ", ".join(values)
            lines.append(f"INSERT INTO {table_name} ({col_list}) VALUES ({values_str});")
            
            # Add blank line every batch_size rows for readability
            if (i + 1) % batch_size == 0:
                lines.append("")
        
        content = "\n".join(lines)
        
        return {
            "success": True,
            "format": "sql",
            "content": content,
            "content_type": "text/plain",
            "filename": f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
            "row_count": len(data),
            "column_count": len(columns)
        }
    
    # ==================== Table Export ====================
    
    def export_table(self, database: str, schema: str, table: str,
                     format: str = 'csv', options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Export a single table.
        
        Args:
            database: Database name
            schema: Schema name
            table: Table name
            format: Export format
            options: Export options
            
        Returns:
            Export result dictionary
        """
        options = options or {}
        full_name = f"{database.upper()}.{schema.upper()}.{table.upper()}"
        
        # Get table info for DDL if needed
        table_info = self.metadata.get_table_info(database.upper(), schema.upper(), table.upper())
        
        if format == 'sql' and options.get('include_ddl', True):
            # Export DDL + data
            return self._export_table_ddl_and_data(database, schema, table, table_info, options)
        else:
            # Export data only
            sql = options.get('query', f"SELECT * FROM {full_name}")
            if options.get('limit'):
                sql += f" LIMIT {options['limit']}"
            
            options['table_name'] = full_name
            return self.export_query_result(sql, format, options)
    
    def _export_table_ddl_and_data(self, database: str, schema: str, table: str,
                                   table_info: Dict, options: Dict[str, Any]) -> Dict[str, Any]:
        """Export table DDL and data as SQL."""
        full_name = f"{database.upper()}.{schema.upper()}.{table.upper()}"
        
        lines = [
            f"-- Export of table {full_name}",
            f"-- Generated by Snowglobe on {datetime.now().isoformat()}",
            ""
        ]
        
        # Generate CREATE TABLE
        if table_info and table_info.get('columns'):
            col_defs = []
            for col in table_info['columns']:
                col_def = f"    {col['name']} {col['type']}"
                if not col.get('nullable', True):
                    col_def += " NOT NULL"
                col_defs.append(col_def)
            
            lines.append(f"CREATE TABLE IF NOT EXISTS {full_name} (")
            lines.append(",\n".join(col_defs))
            lines.append(");")
            lines.append("")
        
        # Get and add data
        result = self.executor.execute(f"SELECT * FROM {full_name}")
        if result["success"] and result["data"]:
            columns = result["columns"]
            data = result["data"]
            
            col_list = ", ".join(columns)
            
            for row in data:
                values = []
                for v in row:
                    if v is None:
                        values.append("NULL")
                    elif isinstance(v, str):
                        escaped = v.replace("'", "''")
                        values.append(f"'{escaped}'")
                    elif isinstance(v, bool):
                        values.append("TRUE" if v else "FALSE")
                    else:
                        values.append(str(v))
                
                values_str = ", ".join(values)
                lines.append(f"INSERT INTO {full_name} ({col_list}) VALUES ({values_str});")
        
        content = "\n".join(lines)
        
        return {
            "success": True,
            "format": "sql",
            "content": content,
            "content_type": "text/plain",
            "filename": f"{table.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
            "row_count": len(result.get("data", [])),
            "table": full_name
        }
    
    # ==================== Schema Export ====================
    
    def export_schema(self, database: str, schema: str, format: str = 'sql',
                      options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Export an entire schema (all tables).
        
        Args:
            database: Database name
            schema: Schema name
            format: Export format
            options: Export options
            
        Returns:
            Export result dictionary (typically a zip file for multiple tables)
        """
        options = options or {}
        include_data = options.get('include_data', True)
        include_views = options.get('include_views', True)
        tables_filter = options.get('tables', None)  # List of table names to include
        
        db_upper = database.upper()
        schema_upper = schema.upper()
        
        # Get all tables in the schema
        tables = self.metadata.list_tables(db_upper, schema_upper)
        views = self.metadata.list_views(db_upper, schema_upper) if include_views else []
        
        if tables_filter:
            tables = [t for t in tables if t['name'].upper() in [n.upper() for n in tables_filter]]
        
        if format == 'sql':
            # Single SQL file with all DDL and data
            return self._export_schema_sql(db_upper, schema_upper, tables, views, options)
        else:
            # Zip file with one file per table
            return self._export_schema_zip(db_upper, schema_upper, tables, format, options)
    
    def _export_schema_sql(self, database: str, schema: str, tables: List[Dict],
                           views: List[Dict], options: Dict[str, Any]) -> Dict[str, Any]:
        """Export schema as a single SQL file."""
        include_data = options.get('include_data', True)
        
        lines = [
            f"-- Schema export: {database}.{schema}",
            f"-- Generated by Snowglobe on {datetime.now().isoformat()}",
            f"-- Tables: {len(tables)}, Views: {len(views)}",
            "",
            f"-- Create schema if not exists",
            f"CREATE SCHEMA IF NOT EXISTS {database}.{schema};",
            f"USE SCHEMA {database}.{schema};",
            ""
        ]
        
        total_rows = 0
        
        # Export tables
        for table in tables:
            table_name = table['name']
            full_name = f"{database}.{schema}.{table_name}"
            
            lines.append(f"-- Table: {table_name}")
            lines.append("-- " + "=" * 50)
            
            # Get columns
            table_info = self.metadata.get_table_info(database, schema, table_name)
            if table_info and table_info.get('columns'):
                col_defs = []
                for col in table_info['columns']:
                    col_def = f"    {col['name']} {col['type']}"
                    if not col.get('nullable', True):
                        col_def += " NOT NULL"
                    col_defs.append(col_def)
                
                lines.append(f"CREATE TABLE IF NOT EXISTS {table_name} (")
                lines.append(",\n".join(col_defs))
                lines.append(");")
                lines.append("")
            
            # Export data if requested
            if include_data:
                result = self.executor.execute(f"SELECT * FROM {full_name}")
                if result["success"] and result["data"]:
                    columns = result["columns"]
                    data = result["data"]
                    total_rows += len(data)
                    
                    col_list = ", ".join(columns)
                    for row in data:
                        values = []
                        for v in row:
                            if v is None:
                                values.append("NULL")
                            elif isinstance(v, str):
                                escaped = v.replace("'", "''")
                                values.append(f"'{escaped}'")
                            elif isinstance(v, bool):
                                values.append("TRUE" if v else "FALSE")
                            else:
                                values.append(str(v))
                        
                        values_str = ", ".join(values)
                        lines.append(f"INSERT INTO {table_name} ({col_list}) VALUES ({values_str});")
            
            lines.append("")
        
        # Export views
        for view in views:
            view_name = view['name']
            lines.append(f"-- View: {view_name}")
            lines.append("-- " + "=" * 50)
            
            definition = view.get('definition', '')
            if definition:
                lines.append(f"CREATE OR REPLACE VIEW {view_name} AS")
                lines.append(definition + ";")
            lines.append("")
        
        content = "\n".join(lines)
        
        return {
            "success": True,
            "format": "sql",
            "content": content,
            "content_type": "text/plain",
            "filename": f"{schema.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
            "table_count": len(tables),
            "view_count": len(views),
            "row_count": total_rows,
            "schema": f"{database}.{schema}"
        }
    
    def _export_schema_zip(self, database: str, schema: str, tables: List[Dict],
                           format: str, options: Dict[str, Any]) -> Dict[str, Any]:
        """Export schema as a zip file with one file per table."""
        output = io.BytesIO()
        
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add metadata file
            metadata = {
                "database": database,
                "schema": schema,
                "tables": [t['name'] for t in tables],
                "exported_at": datetime.now().isoformat(),
                "format": format
            }
            zf.writestr("_metadata.json", json.dumps(metadata, indent=2))
            
            # Export each table
            for table in tables:
                table_name = table['name']
                full_name = f"{database}.{schema}.{table_name}"
                
                # Get data
                result = self.executor.execute(f"SELECT * FROM {full_name}")
                if result["success"]:
                    columns = result["columns"]
                    data = result["data"]
                    
                    export_result = self._export_data(columns, data, format, options)
                    if export_result["success"]:
                        content = export_result["content"]
                        if isinstance(content, str):
                            content = content.encode('utf-8')
                        
                        ext = format if format != 'jsonl' else 'jsonl'
                        zf.writestr(f"{table_name.lower()}.{ext}", content)
        
        output.seek(0)
        
        return {
            "success": True,
            "format": "zip",
            "content": output.getvalue(),
            "content_type": "application/zip",
            "filename": f"{schema.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            "table_count": len(tables),
            "binary": True
        }
    
    # ==================== Database Export ====================
    
    def export_database(self, database: str, format: str = 'sql',
                        options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Export an entire database (all schemas and tables).
        
        Args:
            database: Database name
            format: Export format
            options: Export options
            
        Returns:
            Export result dictionary (typically a zip file)
        """
        options = options or {}
        include_data = options.get('include_data', True)
        schemas_filter = options.get('schemas', None)
        
        db_upper = database.upper()
        
        # Get all schemas
        schemas = self.metadata.list_schemas(db_upper)
        
        if schemas_filter:
            schemas = [s for s in schemas if s['name'].upper() in [n.upper() for n in schemas_filter]]
        
        # Skip system schemas
        skip_schemas = {'INFORMATION_SCHEMA'}
        schemas = [s for s in schemas if s['name'].upper() not in skip_schemas]
        
        if format == 'sql':
            return self._export_database_sql(db_upper, schemas, options)
        else:
            return self._export_database_zip(db_upper, schemas, format, options)
    
    def _export_database_sql(self, database: str, schemas: List[Dict],
                             options: Dict[str, Any]) -> Dict[str, Any]:
        """Export database as a single SQL file."""
        include_data = options.get('include_data', True)
        
        lines = [
            f"-- Database export: {database}",
            f"-- Generated by Snowglobe on {datetime.now().isoformat()}",
            f"-- Schemas: {len(schemas)}",
            "",
            f"-- Create database if not exists",
            f"CREATE DATABASE IF NOT EXISTS {database};",
            f"USE DATABASE {database};",
            ""
        ]
        
        total_tables = 0
        total_rows = 0
        
        for schema_info in schemas:
            schema_name = schema_info['name']
            
            lines.append(f"-- Schema: {schema_name}")
            lines.append("-- " + "=" * 60)
            lines.append(f"CREATE SCHEMA IF NOT EXISTS {schema_name};")
            lines.append(f"USE SCHEMA {schema_name};")
            lines.append("")
            
            # Get tables
            tables = self.metadata.list_tables(database, schema_name)
            total_tables += len(tables)
            
            for table in tables:
                table_name = table['name']
                full_name = f"{database}.{schema_name}.{table_name}"
                
                lines.append(f"-- Table: {table_name}")
                
                # Get columns
                table_info = self.metadata.get_table_info(database, schema_name, table_name)
                if table_info and table_info.get('columns'):
                    col_defs = []
                    for col in table_info['columns']:
                        col_def = f"    {col['name']} {col['type']}"
                        if not col.get('nullable', True):
                            col_def += " NOT NULL"
                        col_defs.append(col_def)
                    
                    lines.append(f"CREATE TABLE IF NOT EXISTS {table_name} (")
                    lines.append(",\n".join(col_defs))
                    lines.append(");")
                
                # Export data
                if include_data:
                    result = self.executor.execute(f"SELECT * FROM {full_name}")
                    if result["success"] and result["data"]:
                        columns = result["columns"]
                        data = result["data"]
                        total_rows += len(data)
                        
                        col_list = ", ".join(columns)
                        for row in data:
                            values = []
                            for v in row:
                                if v is None:
                                    values.append("NULL")
                                elif isinstance(v, str):
                                    escaped = v.replace("'", "''")
                                    values.append(f"'{escaped}'")
                                elif isinstance(v, bool):
                                    values.append("TRUE" if v else "FALSE")
                                else:
                                    values.append(str(v))
                            
                            values_str = ", ".join(values)
                            lines.append(f"INSERT INTO {table_name} ({col_list}) VALUES ({values_str});")
                
                lines.append("")
            
            # Export views
            views = self.metadata.list_views(database, schema_name)
            for view in views:
                view_name = view['name']
                definition = view.get('definition', '')
                if definition:
                    lines.append(f"-- View: {view_name}")
                    lines.append(f"CREATE OR REPLACE VIEW {view_name} AS")
                    lines.append(definition + ";")
                    lines.append("")
        
        content = "\n".join(lines)
        
        return {
            "success": True,
            "format": "sql",
            "content": content,
            "content_type": "text/plain",
            "filename": f"{database.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
            "schema_count": len(schemas),
            "table_count": total_tables,
            "row_count": total_rows,
            "database": database
        }
    
    def _export_database_zip(self, database: str, schemas: List[Dict],
                             format: str, options: Dict[str, Any]) -> Dict[str, Any]:
        """Export database as a zip file with folders per schema."""
        output = io.BytesIO()
        
        total_tables = 0
        
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add metadata
            metadata = {
                "database": database,
                "schemas": [s['name'] for s in schemas],
                "exported_at": datetime.now().isoformat(),
                "format": format
            }
            zf.writestr("_metadata.json", json.dumps(metadata, indent=2))
            
            for schema_info in schemas:
                schema_name = schema_info['name']
                tables = self.metadata.list_tables(database, schema_name)
                total_tables += len(tables)
                
                for table in tables:
                    table_name = table['name']
                    full_name = f"{database}.{schema_name}.{table_name}"
                    
                    result = self.executor.execute(f"SELECT * FROM {full_name}")
                    if result["success"]:
                        columns = result["columns"]
                        data = result["data"]
                        
                        export_result = self._export_data(columns, data, format, options)
                        if export_result["success"]:
                            content = export_result["content"]
                            if isinstance(content, str):
                                content = content.encode('utf-8')
                            
                            ext = format if format != 'jsonl' else 'jsonl'
                            file_path = f"{schema_name.lower()}/{table_name.lower()}.{ext}"
                            zf.writestr(file_path, content)
        
        output.seek(0)
        
        return {
            "success": True,
            "format": "zip",
            "content": output.getvalue(),
            "content_type": "application/zip",
            "filename": f"{database.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            "schema_count": len(schemas),
            "table_count": total_tables,
            "binary": True
        }
    
    # ==================== DDL Export ====================
    
    def export_ddl(self, database: str = None, schema: str = None,
                   object_type: str = None) -> Dict[str, Any]:
        """
        Export DDL statements for database objects.
        
        Args:
            database: Database name (all if None)
            schema: Schema name (all in database if None)
            object_type: Type of objects (TABLE, VIEW, all if None)
            
        Returns:
            Dictionary with DDL content
        """
        lines = [
            "-- DDL Export",
            f"-- Generated by Snowglobe on {datetime.now().isoformat()}",
            ""
        ]
        
        if database:
            databases = [{'name': database.upper()}]
        else:
            databases = self.metadata.list_databases()
        
        for db in databases:
            db_name = db['name']
            lines.append(f"-- Database: {db_name}")
            lines.append(f"CREATE DATABASE IF NOT EXISTS {db_name};")
            lines.append(f"USE DATABASE {db_name};")
            lines.append("")
            
            if schema:
                schemas = [{'name': schema.upper()}]
            else:
                schemas = self.metadata.list_schemas(db_name)
                schemas = [s for s in schemas if s['name'] != 'INFORMATION_SCHEMA']
            
            for sch in schemas:
                schema_name = sch['name']
                lines.append(f"-- Schema: {schema_name}")
                lines.append(f"CREATE SCHEMA IF NOT EXISTS {schema_name};")
                lines.append("")
                
                # Tables
                if not object_type or object_type.upper() == 'TABLE':
                    tables = self.metadata.list_tables(db_name, schema_name)
                    for table in tables:
                        table_name = table['name']
                        table_info = self.metadata.get_table_info(db_name, schema_name, table_name)
                        
                        if table_info and table_info.get('columns'):
                            col_defs = []
                            for col in table_info['columns']:
                                col_def = f"    {col['name']} {col['type']}"
                                if not col.get('nullable', True):
                                    col_def += " NOT NULL"
                                col_defs.append(col_def)
                            
                            lines.append(f"CREATE TABLE IF NOT EXISTS {schema_name}.{table_name} (")
                            lines.append(",\n".join(col_defs))
                            lines.append(");")
                            lines.append("")
                
                # Views
                if not object_type or object_type.upper() == 'VIEW':
                    views = self.metadata.list_views(db_name, schema_name)
                    for view in views:
                        view_name = view['name']
                        definition = view.get('definition', '')
                        if definition:
                            lines.append(f"CREATE OR REPLACE VIEW {schema_name}.{view_name} AS")
                            lines.append(definition + ";")
                            lines.append("")
        
        content = "\n".join(lines)
        
        return {
            "success": True,
            "format": "sql",
            "content": content,
            "content_type": "text/plain",
            "filename": f"ddl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
        }
    
    # ==================== Selective Export ====================
    
    def export_tables(self, tables: List[Dict[str, str]], format: str = 'csv',
                      options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Export multiple tables by specification.
        
        Args:
            tables: List of table specs [{"database": "", "schema": "", "table": ""}]
            format: Export format
            options: Export options
            
        Returns:
            Export result (zip file)
        """
        options = options or {}
        output = io.BytesIO()
        
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add metadata
            metadata = {
                "tables": tables,
                "exported_at": datetime.now().isoformat(),
                "format": format
            }
            zf.writestr("_metadata.json", json.dumps(metadata, indent=2))
            
            for table_spec in tables:
                database = table_spec.get('database', '').upper()
                schema = table_spec.get('schema', '').upper()
                table = table_spec.get('table', '').upper()
                
                if not all([database, schema, table]):
                    continue
                
                full_name = f"{database}.{schema}.{table}"
                
                result = self.executor.execute(f"SELECT * FROM {full_name}")
                if result["success"]:
                    columns = result["columns"]
                    data = result["data"]
                    
                    export_result = self._export_data(columns, data, format, options)
                    if export_result["success"]:
                        content = export_result["content"]
                        if isinstance(content, str):
                            content = content.encode('utf-8')
                        
                        ext = format if format != 'jsonl' else 'jsonl'
                        file_path = f"{database.lower()}/{schema.lower()}/{table.lower()}.{ext}"
                        zf.writestr(file_path, content)
        
        output.seek(0)
        
        return {
            "success": True,
            "format": "zip",
            "content": output.getvalue(),
            "content_type": "application/zip",
            "filename": f"tables_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            "table_count": len(tables),
            "binary": True
        }


# Export format configurations
EXPORT_FORMATS = {
    "csv": {
        "name": "CSV",
        "extension": ".csv",
        "content_type": "text/csv",
        "description": "Comma-separated values",
        "options": ["delimiter", "quote_char", "include_header", "null_value"]
    },
    "json": {
        "name": "JSON",
        "extension": ".json",
        "content_type": "application/json",
        "description": "JSON format",
        "options": ["orient", "indent"]
    },
    "jsonl": {
        "name": "JSON Lines",
        "extension": ".jsonl",
        "content_type": "application/x-ndjson",
        "description": "One JSON object per line",
        "options": []
    },
    "parquet": {
        "name": "Parquet",
        "extension": ".parquet",
        "content_type": "application/octet-stream",
        "description": "Apache Parquet columnar format",
        "options": ["compression"],
        "requires": ["pyarrow"]
    },
    "excel": {
        "name": "Excel",
        "extension": ".xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "description": "Microsoft Excel format",
        "options": ["sheet_name"],
        "requires": ["openpyxl"]
    },
    "sql": {
        "name": "SQL",
        "extension": ".sql",
        "content_type": "text/plain",
        "description": "SQL INSERT statements",
        "options": ["table_name", "include_create", "batch_size"]
    }
}
